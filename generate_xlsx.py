import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

REPORT = r'c:\Users\phucl\OneDrive\Desktop\phuc\TESTING\project2\report.md'
TEMPLATE = r'c:\Users\phucl\OneDrive\Desktop\phuc\TESTING\project2\Testcase-template.xlsx'
OUTPUT = r'c:\Users\phucl\OneDrive\Desktop\phuc\TESTING\project2\Testcase-filled.xlsx'

FEATURE_INFO = {
    '001': {
        'name': 'Admin Adds a New User',
        'desc': 'Verify admin can create a user account with valid credentials and that invalid inputs are rejected.',
        'tester': '[Member 1]',
        'precond': 'Log in to https://ihatetesting.moodlecloud.com/ as Manager. Navigate to: Site administration > Users > Accounts > Add a new user.',
    },
    '002': {
        'name': 'Admin Creates a New Course',
        'desc': 'Verify admin can create a course with unique short name and valid date range.',
        'tester': '[Member 1]',
        'precond': 'Log in as Manager. Navigate to: Site administration > Courses > Manage courses and categories. Click "Create new course".',
    },
    '003': {
        'name': 'Teacher Creates Assignment',
        'desc': 'Verify teacher can create an assignment with valid name, grade limits, and submission dates.',
        'tester': '[Member 2]',
        'precond': 'Log in as Teacher. Open the target course. Turn Edit mode ON (toggle, top-right). Click "Add an activity or resource" and select "Assignment".',
    },
    '004': {
        'name': 'Teacher Grades an Assignment',
        'desc': 'Verify teacher can enter a numeric grade (0-100) and that out-of-range values are rejected.',
        'tester': '[Member 2]',
        'precond': 'Log in as Teacher. Open the target course > Assignment. Click "View all submissions". Click the Grade icon (pencil) for the target student.',
    },
    '005': {
        'name': 'User Creates Calendar Event',
        'desc': 'Verify users can create calendar events with valid title and optional duration settings.',
        'tester': '[Member 3]',
        'precond': 'Log in to Moodle (any role). Click "Calendar" from the Dashboard or side navigation. Click "New event".',
    },
    '006': {
        'name': 'Teacher Sets Up a Quiz',
        'desc': 'Verify teacher can create a quiz with valid grade limits, time settings, and open/close dates.',
        'tester': '[Member 4]',
        'precond': 'Log in as Teacher. Open the target course. Turn Edit mode ON. Click "Add an activity or resource" and select "Quiz".',
    },
}

def clean(v):
    return v.strip().strip('`').strip()

def parse_tcs(report_path):
    """Parse all TC rows from report.md. Returns dict: feat_id -> list of row dicts."""
    with open(report_path, encoding='utf-8') as f:
        content = f.read()
    # Split into feature sections
    feat_matches = list(re.finditer(r'^# Feature (\d{3}):', content, re.MULTILINE))
    all_tcs = {}
    for idx, m in enumerate(feat_matches):
        fid = m.group(1)
        start = m.start()
        end = feat_matches[idx+1].start() if idx+1 < len(feat_matches) else len(content)
        section = content[start:end]
        tcs = []
        lines = section.splitlines()
        i = 0
        while i < len(lines):
            line = lines[i]
            # detect header row of a TC table
            if re.search(r'\|\s*TC ID\s*\|', line):
                headers = [h.strip() for h in line.split('|') if h.strip()]
                i += 1  # skip separator
                while i < len(lines) and re.match(r'\|[\s\-|]+\|', lines[i]):
                    i += 1
                # read data rows
                while i < len(lines) and lines[i].strip().startswith('|'):
                    dline = lines[i]
                    if re.match(r'\|[\s\-|]+\|', dline):
                        i += 1; continue
                    cols = [c.strip() for c in dline.split('|')]
                    cols = [c for c in cols if c != '' or len(cols) > 2]
                    # strip leading/trailing empty from split
                    while cols and cols[0] == '': cols.pop(0)
                    while cols and cols[-1] == '': cols.pop()
                    if cols and re.match(r'TC-\d{3}-\d{3}', cols[0]):
                        entry = {}
                        for j, h in enumerate(headers):
                            entry[h] = clean(cols[j]) if j < len(cols) else ''
                        tcs.append(entry)
                    i += 1
                continue
            i += 1
        all_tcs[fid] = tcs
    return all_tcs

def tc_name_desc(tc, fid):
    """Build [TECHNIQUE] name and description for a TC."""
    tech = tc.get('Technique', tc.get('technique', 'BVA')).strip()
    prefix = f'[{tech}]'
    raw_name = tc.get('Variable Tested', tc.get('Test Case Name', tc.get('Test case name', ''))).strip()

    if tech == 'BVA':
        name = f'{prefix} {raw_name}'
        desc = f'BVA (Robust Single Fault): verify that the boundary value condition "{raw_name}" produces the expected result while all other variables are held at nominal values.'
    elif tech == 'ECP':
        name = f'{prefix} {raw_name}'
        desc = f'ECP (Weak Robust): verify that the equivalence class "{raw_name}" produces the expected result while all other variables are held at nominal values.'
    elif tech == 'UC':
        name = f'{prefix} {raw_name}'
        desc = f'Use-Case testing: verify the scenario "{raw_name}" traces the correct path through the activity diagram.'
    elif tech == 'DT':
        name = f'{prefix} {raw_name}'
        desc = f'Decision Table testing: verify that the combination of conditions in rule "{raw_name}" produces the correct system action.'
    else:
        name = f'{prefix} {raw_name}'
        desc = f'{tech} test: verify "{raw_name}".'
    return name, desc

def gen_steps(tc, fid):
    """Generate (steps_list, submit_step, verify_note, expected_result) for a TC."""
    exp = tc.get('Expected Result', '')
    is_pass = '✅' in exp

    def field_step(label, field_key, action_pattern, disabled_text=None):
        v = tc.get(field_key, '')
        if not v:
            return None
        if v == '*(empty)*':
            return f'Leave the "{label}" field empty (do not type anything).'
        if disabled_text and v.lower() == 'disabled':
            return disabled_text
        return action_pattern.format(v)

    steps = []
    if fid == '001':
        u = tc.get('Username', '')
        p = tc.get('Password', '')
        f = tc.get('First Name', '')
        l = tc.get('Last Name', '')
        e = tc.get('Email', '')
        steps.append('Leave the "Username" field empty.' if u == '*(empty)*' else f'In the "Username" field, type: {u}')
        if p == '*(empty)*':
            steps.append('Check the "Generate password and notify user" checkbox. Leave the password field empty.')
        else:
            steps.append(f'In the "New password" field, type: {p}')
        steps.append('Leave the "First name" field empty.' if f == '*(empty)*' else f'In the "First name" field, type: {f}')
        steps.append('Leave the "Last name" field empty.' if l == '*(empty)*' else f'In the "Last name" field, type: {l}')
        steps.append('Leave the "Email address" field empty.' if e == '*(empty)*' else f'In the "Email address" field, type: {e}')
        submit = 'Click the "Create user" button.'
        if is_pass:
            verify = f'Verify: The page redirects to the user list and the new user "{u}" appears in the list. No error messages are shown.'
        else:
            verify = 'Verify: An error message appears on the form (e.g., near the Username, Password, Name, or Email field). The page does not redirect and no user is created.'
        result = '✅ User created successfully.' if is_pass else '❌ Form submission fails. Error message shown on form.'

    elif fid == '002':
        fn = tc.get('Full Name', '')
        sn = tc.get('Short Name', '')
        ed = tc.get('End Date', '')
        sec = tc.get('Sections', '')
        steps.append('Leave the "Course full name" field empty.' if fn == '*(empty)*' else f'In the "Course full name" field, type: {fn}')
        steps.append('Leave the "Course short name" field empty.' if sn == '*(empty)*' else f'In the "Course short name" field, type: {sn}')
        if ed.lower() == 'disabled':
            steps.append('Leave the end date "Enable" checkbox unchecked (end date disabled).')
        else:
            steps.append(f'Check the "Enable" checkbox for "Course end date". Set the end date to: {ed} (relative to the start date set to today).')
        steps.append(f'Set "Number of sections" to: {sec}')
        submit = 'Click "Save and display".'
        if is_pass:
            verify = f'Verify: The course page opens and the course title "{fn}" appears in the page header. No error messages are shown.'
        else:
            verify = 'Verify: An error message appears on the form (e.g., near the Full Name, Short Name, or End Date field). The course is not created.'
        result = '✅ Course created successfully.' if is_pass else '❌ Course creation fails. Error message shown on form.'

    elif fid == '003':
        an = tc.get('Assig. Name', '')
        gp = tc.get('Grade to Pass', '')
        dd = tc.get('Due Date', '')
        cd = tc.get('Cut-off Date', '')
        steps.append('Leave the "Assignment name" field empty.' if an == '*(empty)*' else f'In the "Assignment name" field, type: {an}')
        steps.append('Check "Enable" for "Allow submissions from". Set it to today\'s date.')
        if dd.lower() == 'disabled':
            steps.append('Leave the "Due date" Enable checkbox unchecked (due date disabled).')
        else:
            steps.append(f'Check "Enable" for "Due date". Set it to: {dd} (relative to today\'s Allow date).')
        if cd.lower() == 'disabled':
            steps.append('Leave the "Cut-off date" Enable checkbox unchecked (cut-off disabled).')
        else:
            steps.append(f'Check "Enable" for "Cut-off date". Set it to: {cd} (relative to the Due date).')
        steps.append(f'In the "Grade to pass" field, type: {gp}')
        submit = 'Click "Save and return to course".'
        if is_pass:
            verify = f'Verify: The assignment "{an}" appears on the course page. No error messages are shown.'
        else:
            verify = 'Verify: An error message appears on the form (e.g., near the Assignment Name, Grade to Pass, or date fields). The assignment is not created.'
        result = '✅ Assignment created successfully.' if is_pass else '❌ Assignment creation fails. Error message shown on form.'

    elif fid == '004':
        gv = tc.get('Grade Value', '')
        fb = tc.get('Feedback', 'Good work')
        nt = tc.get('Notify', 'Yes')
        steps.append('Clear the "Grade" field and leave it empty.' if gv == '*(empty)*' else f'In the "Grade" field, clear any existing value and type: {gv}')
        steps.append(f'In the "Feedback comments" rich text editor, type: {fb}')
        steps.append(f'Set "Notify student" checkbox to: {nt}')
        submit = 'Click "Save changes".'
        if is_pass:
            verify = f'Verify: The grade "{gv}" is saved and shown in the grading panel. A confirmation message or the next student view is displayed.'
        else:
            verify = f'Verify: An error message appears near the Grade field indicating the value "{gv}" is invalid or out of range. The grade is not saved.'
        result = '✅ Grade saved successfully.' if is_pass else '❌ Grade save fails. Error message shown near Grade field.'

    elif fid == '005':
        tt = tc.get('Event Title', '')
        dm = tc.get('Duration Min', '')
        ud = tc.get('Until Date', '')
        steps.append('Leave the "Event title" field empty.' if tt == '*(empty)*' else f'In the "Event title" field, type: {tt}')
        steps.append('Set the event start date to today\'s date using the date picker.')
        if dm.lower() == 'disabled':
            steps.append('Select "Without duration" as the duration option.')
        else:
            steps.append(f'Select "Duration in minutes" as the duration option. In the minutes field, type: {dm}')
        if ud.lower() == 'disabled':
            steps.append('(No "Until date" set — skip this step as duration is not "Until date".)')
        else:
            steps.append(f'Select "Until date" as the duration option. Set the until date to: {ud} (relative to today\'s event date).')
        submit = 'Click "Save".'
        if is_pass:
            verify = f'Verify: The event "{tt}" appears on the calendar for today\'s date. No error messages are shown.'
        else:
            verify = 'Verify: An error message appears on the form (e.g., near the Event Title or Duration field). The event is not saved.'
        result = '✅ Event created on calendar.' if is_pass else '❌ Event creation fails. Error message shown on form.'

    elif fid == '006':
        qn = tc.get('Quiz Name', '')
        gp = tc.get('Grade Pass', '')
        tl = tc.get('Time Limit', '')
        cd = tc.get('Close Date', '')
        steps.append('Leave the "Name" field empty.' if qn == '*(empty)*' else f'In the "Name" field, type: {qn}')
        steps.append('Check "Enable" for "Open the quiz". Set the open date to today\'s date.')
        if cd.lower() == 'disabled':
            steps.append('Leave the "Close the quiz" Enable checkbox unchecked (close date disabled).')
        else:
            steps.append(f'Check "Enable" for "Close the quiz". Set the close date to: {cd} (relative to today\'s open date).')
        if tl.lower() == 'disabled':
            steps.append('Leave the "Time limit" Enable checkbox unchecked (time limit disabled).')
        else:
            steps.append(f'Check "Enable" for "Time limit". In the time limit field, type: {tl} (unit: minutes).')
        steps.append(f'In the "Grade to pass" field, type: {gp}')
        submit = 'Click "Save and return to course".'
        if is_pass:
            verify = f'Verify: The quiz "{qn}" appears on the course page. No error messages are shown.'
        else:
            verify = 'Verify: An error message appears on the form (e.g., near the Name, Grade to Pass, Time Limit, or Close Date field). The quiz is not created.'
        result = '✅ Quiz created successfully.' if is_pass else '❌ Quiz creation fails. Error message shown on form.'
    else:
        steps = ['Perform the test action as described.']
        submit = 'Click the submit/save button.'
        verify = 'Verify the expected result matches.'
        result = exp

    return steps, submit, verify, result

# ---- STYLES ----
def hdr_style(cell, bold=True, bg=None, wrap=True, halign='center', valign='center'):
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)

def data_style(cell, wrap=True, valign='top', halign='left', bold=False):
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)

thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_border(cell):
    cell.border = BORDER

COLORS = {
    'feat_hdr': 'C6EFCE',   # green — feature group row
    'col_hdr':  'BDD7EE',   # blue — column headers
    'bva':      'EBF0FF',
    'ecp':      'FFF2CC',
    'uc':       'FCE4D6',
    'dt':       'E2EFDA',
    'last_step':'F4F4F4',
}

TC_COLORS = {'BVA': 'EBF0FF', 'ECP': 'FFF2CC', 'UC': 'FCE4D6', 'DT': 'E2EFDA'}

def build_feature_sheet(ws, fid, tcs):
    info = FEATURE_INFO[fid]
    # --- header rows ---
    ws.append(['Product name:', 'Mount Orange University Moodle (https://ihatetesting.moodlecloud.com/)', '', '', '', '', '', ''])
    ws.merge_cells(f'B1:H1')
    ws.append(['Tester:', info['tester'], '', '', 'Last updated:', '', '', ''])
    ws.merge_cells(f'B2:D2')
    ws.merge_cells(f'F2:H2')

    # column headers
    col_hdrs = ['Test case ID', 'Test case name', 'Test case description', 'Precondition',
                'Step name', 'Step description', 'Expected Result', 'Note']
    ws.append(col_hdrs)
    for c in range(1, 9):
        cell = ws.cell(row=3, column=c)
        hdr_style(cell, bg=COLORS['col_hdr'])
        apply_border(cell)

    # feature group row
    ws.append([f'Feature {fid}: {info["name"]}', '', '', '', '', '', '', ''])
    ws.merge_cells(f'A4:H4')
    feat_cell = ws.cell(row=4, column=1)
    hdr_style(feat_cell, bg=COLORS['feat_hdr'], halign='left')
    apply_border(feat_cell)

    cur_row = 5
    for tc in tcs:
        tc_id = tc.get('TC ID', tc.get('Test case ID', ''))
        tech = tc.get('Technique', 'BVA')
        name, desc = tc_name_desc(tc, fid)
        precond = info['precond']
        steps, submit, verify, result = gen_steps(tc, fid)
        all_steps = [(f'Step {i+1}', s) for i, s in enumerate(steps)] + [(f'Step {len(steps)+1}', submit)]
        num_rows = len(all_steps)
        tc_color = TC_COLORS.get(tech, 'FFFFFF')

        # Write TC meta in col A-D, merged vertically
        def write_meta(col, val, bold=False):
            ws.cell(row=cur_row, column=col, value=val)
            if num_rows > 1:
                ws.merge_cells(start_row=cur_row, start_column=col, end_row=cur_row+num_rows-1, end_column=col)
            c = ws.cell(row=cur_row, column=col)
            data_style(c, valign='top', bold=bold)
            c.fill = PatternFill('solid', fgColor=tc_color)
            for r in range(cur_row, cur_row+num_rows):
                apply_border(ws.cell(row=r, column=col))

        write_meta(1, tc_id, bold=True)
        write_meta(2, name)
        write_meta(3, desc)
        write_meta(4, precond)

        # Write step rows
        for s_idx, (step_name, step_desc) in enumerate(all_steps):
            r = cur_row + s_idx
            is_last = (s_idx == num_rows - 1)
            ws.cell(row=r, column=5, value=step_name)
            ws.cell(row=r, column=6, value=step_desc)
            ws.cell(row=r, column=7, value=result if is_last else '')
            ws.cell(row=r, column=8, value=verify if is_last else '')
            for c in range(5, 9):
                cell = ws.cell(row=r, column=c)
                bg = COLORS['last_step'] if is_last else 'FFFFFF'
                cell.fill = PatternFill('solid', fgColor=bg)
                data_style(cell, valign='top')
                apply_border(cell)

        cur_row += num_rows

    # column widths
    widths = [15, 35, 50, 45, 10, 55, 30, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def build_summary_sheet(ws, all_tcs):
    headers = ['Test case ID', 'Test case name', 'Test case description']
    ws.append(headers)
    for c in range(1, 4):
        cell = ws.cell(row=1, column=c)
        hdr_style(cell, bg=COLORS['col_hdr'])
        apply_border(cell)
    r = 2
    for fid in sorted(all_tcs.keys()):
        info = FEATURE_INFO[fid]
        ws.cell(row=r, column=1, value=f'Feature {fid}: {info["name"]}')
        ws.merge_cells(f'A{r}:C{r}')
        gc = ws.cell(row=r, column=1)
        hdr_style(gc, bg=COLORS['feat_hdr'], halign='left')
        apply_border(gc)
        r += 1
        for tc in all_tcs[fid]:
            tc_id = tc.get('TC ID', '')
            name, desc = tc_name_desc(tc, fid)
            ws.cell(row=r, column=1, value=tc_id)
            ws.cell(row=r, column=2, value=name)
            ws.cell(row=r, column=3, value=desc)
            for c in range(1, 4):
                cell = ws.cell(row=r, column=c)
                data_style(cell)
                apply_border(cell)
            r += 1
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 70

def build_project_summary(ws, all_tcs):
    # Find existing data rows and fill them
    # Template has rows 1-4 as headers, row 7 as feature header, rows 8-9 as sample features
    # We'll overwrite from row 8 onwards
    start_row = 8
    for i, fid in enumerate(sorted(all_tcs.keys())):
        info = FEATURE_INFO[fid]
        r = start_row + i
        ws.cell(row=r, column=1, value=fid)
        ws.cell(row=r, column=2, value=info['name'])
        ws.cell(row=r, column=3, value='5.2')
        ws.cell(row=r, column=4, value=info['desc'])
        ws.cell(row=r, column=5, value=info['tester'])

def main():
    print('Parsing report.md...')
    all_tcs = parse_tcs(REPORT)
    for fid, tcs in all_tcs.items():
        print(f'  Feature {fid}: {len(tcs)} TCs parsed')

    print('Loading template...')
    wb = load_workbook(TEMPLATE)

    # Fill project summary (Sheet 2)
    print('Filling Test project summary...')
    build_project_summary(wb['Test project summary'], all_tcs)

    # Fill test case summary (Sheet 3)
    print('Filling Test case summary...')
    ws_sum = wb['Test case summary']
    # Unmerge all merged ranges first, then clear
    for mc in list(ws_sum.merged_cells.ranges):
        ws_sum.unmerge_cells(str(mc))
    for r in range(2, ws_sum.max_row + 1):
        for c in range(1, 4):
            ws_sum.cell(row=r, column=c).value = None
    build_summary_sheet(ws_sum, all_tcs)

    # Create feature detail sheets
    template_feat_sheet = wb['<Feature Name 1>']
    for fid in sorted(all_tcs.keys()):
        sheet_name = f'Feature {fid}'
        print(f'Building {sheet_name}...')
        # Create a new sheet for this feature
        ws_new = wb.copy_worksheet(template_feat_sheet)
        ws_new.title = sheet_name
        # Clear all content from copied sheet
        for row in ws_new.iter_rows():
            for cell in row:
                cell.value = None
        # Unmerge all merged cells
        for mc in list(ws_new.merged_cells.ranges):
            ws_new.unmerge_cells(str(mc))
        build_feature_sheet(ws_new, fid, all_tcs[fid])

    # Remove original template sheet
    del wb['<Feature Name 1>']

    print(f'Saving to {OUTPUT}...')
    wb.save(OUTPUT)
    print('Done!')

if __name__ == '__main__':
    main()
